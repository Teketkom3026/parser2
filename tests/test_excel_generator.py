"""Тесты генерации Excel."""

import os
import tempfile

from backend.output.excel_generator import generate_excel


def test_generate_excel_creates_file():
    contacts = [
        {
            "company_name": "ООО Тест",
            "site_url": "test.ru",
            "company_email": "info@test.ru",
            "company_phone": "+7 (495) 111-22-33",
            "person_name": "Иванов Иван",
            "position_raw": "Генеральный директор",
            "position_norm": "Генеральный директор",
            "role_category": "Топ-менеджмент",
            "person_email": "ivanov@test.ru",
            "person_phone": "",
            "inn": "7707083893",
            "kpp": "773601001",
            "social_links": "https://linkedin.com/in/ivanov",
            "page_url": "https://test.ru/team",
            "page_language": "ru",
            "scan_date": "2025-01-15",
            "status": "OK",
            "comment": "",
        }
    ]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    try:
        generate_excel(contacts, path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 1000  # не пустой

        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        assert ws.cell(1, 1).value == "Название компании"
        assert ws.cell(2, 1).value == "ООО Тест"
        assert ws.cell(2, 5).value == "Иванов Иван"

        # Проверяем лист сводки
        assert "Сводка" in wb.sheetnames
        wb.close()
    finally:
        os.unlink(path)


def test_generate_excel_empty():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name

    try:
        generate_excel([], path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 500
    finally:
        os.unlink(path)
