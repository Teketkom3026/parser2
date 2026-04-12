"""Парсинг входных файлов (Excel, CSV) со списком URL."""

import csv
import io
from pathlib import Path

from openpyxl import load_workbook

from backend.utils.url_normalizer import normalize_url, is_valid_url
from backend.utils.logger import get_logger

logger = get_logger("file_parser")


def parse_file(file_path: str) -> list[str]:
    """Парсит файл и возвращает список URL."""
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        urls = _parse_excel(path)
    elif suffix == ".csv":
        urls = _parse_csv(path)
    else:
        raise ValueError(f"Неподдерживаемый формат файла: {suffix}")

    # Нормализация и валидация
    valid_urls = []
    seen: set[str] = set()

    for url in urls:
        normalized = normalize_url(url)
        if normalized and is_valid_url(normalized):
            domain_key = normalized.lower().rstrip("/")
            if domain_key not in seen:
                seen.add(domain_key)
                valid_urls.append(normalized)

    logger.info("file_parsed", path=str(path), raw=len(urls), valid=len(valid_urls))
    return valid_urls


def _parse_excel(path: Path) -> list[str]:
    """Парсит Excel-файл."""
    urls = []

    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    value = str(cell).strip()
                    if _looks_like_url(value):
                        urls.append(value)

        wb.close()

    except Exception as e:
        logger.error("excel_parse_error", path=str(path), error=str(e)[:200])
        raise ValueError(f"Ошибка чтения Excel: {e}")

    return urls


def _parse_csv(path: Path) -> list[str]:
    """Парсит CSV-файл."""
    urls = []

    try:
        with open(path, encoding="utf-8", errors="ignore") as f:
            content = f.read()

        # Пробуем определить разделитель
        try:
            dialect = csv.Sniffer().sniff(content[:2000])
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","

        reader = csv.reader(io.StringIO(content), delimiter=delimiter)

        for row in reader:
            for cell in row:
                value = cell.strip()
                if _looks_like_url(value):
                    urls.append(value)

    except Exception as e:
        logger.error("csv_parse_error", path=str(path), error=str(e)[:200])
        raise ValueError(f"Ошибка чтения CSV: {e}")

    return urls


def parse_urls_text(text: str) -> list[str]:
    """Парсит URL из текстовой строки (через запятую или перенос строки)."""
    if not text:
        return []

    # Разделители
    urls_raw = []
    for sep in ["\n", ",", ";", " "]:
        if sep in text:
            urls_raw = [u.strip() for u in text.split(sep) if u.strip()]
            break

    if not urls_raw:
        urls_raw = [text.strip()]

    valid = []
    seen: set[str] = set()

    for url in urls_raw:
        normalized = normalize_url(url)
        if normalized and is_valid_url(normalized):
            key = normalized.lower().rstrip("/")
            if key not in seen:
                seen.add(key)
                valid.append(normalized)

    return valid


def _looks_like_url(value: str) -> bool:
    """Быстрая проверка — похоже ли на URL."""
    if not value or len(value) < 4:
        return False

    value_lower = value.lower().strip()

    # Явные URL
    if value_lower.startswith(("http://", "https://", "www.")):
        return True

    # Похоже на домен
    if "." in value_lower and " " not in value_lower:
        parts = value_lower.split(".")
        tld = parts[-1]
        if len(tld) >= 2 and len(tld) <= 10 and tld.isalpha():
            return True

    return False
