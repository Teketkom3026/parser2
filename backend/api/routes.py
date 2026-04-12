"""REST API — основные маршруты."""
import io
from pathlib import Path

import openpyxl
from fastapi import APIRouter, File, Form, UploadFile, HTTPException
from fastapi.responses import FileResponse

from backend.utils.logger import get_logger
from backend.utils.url_normalizer import normalize_url

logger = get_logger("api")
router = APIRouter(prefix="/api/v1")


def _get_db():
    from backend.main import app_state
    return app_state["db"]


def _get_tm():
    from backend.main import app_state
    return app_state["task_manager"]


@router.post("/tasks")
async def create_task(
    file: UploadFile = File(...),
    mode: str = Form("mode_2"),
    positions: str = Form(""),
):
    """Создать задачу: загрузить Excel с URL."""
    db = _get_db()
    tm = _get_tm()

    if mode not in ("mode_1", "mode_2"):
        raise HTTPException(400, "mode должен быть mode_1 или mode_2")
    if mode == "mode_1" and not positions.strip():
        raise HTTPException(400, "Для mode_1 нужно указать целевые должности")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        urls = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            for cell in row:
                if cell and str(cell).strip():
                    val = str(cell).strip()
                    if "." in val or "://" in val:
                        urls.append(normalize_url(val))
        wb.close()
    except Exception as e:
        raise HTTPException(400, f"Ошибка чтения Excel: {e}")

    if not urls:
        raise HTTPException(400, "В файле не найдено URL")

    urls = list(dict.fromkeys(urls))

    target_positions = []
    if positions.strip():
        target_positions = [p.strip() for p in positions.split(",") if p.strip()]

    task_id = await tm.create_task(
        urls=urls,
        mode=mode,
        target_positions=target_positions if target_positions else None,
        input_file=file.filename,
    )
    await tm.start_task(task_id)
    return {"task_id": task_id, "status": "running", "total_urls": len(urls)}


@router.get("/tasks")
async def list_tasks():
    db = _get_db()
    tasks = await db.list_tasks(limit=50)
    return [
        {
            "task_id": t["id"],
            "mode": t["mode"],
            "status": t["status"],
            "total_urls": t["total_urls"],
            "processed_urls": t["processed_urls"] or 0,
            "found_contacts": t["found_contacts"] or 0,
            "errors_count": t["errors_count"] or 0,
            "created_at": t["created_at"],
            "output_file": t.get("output_file"),
        }
        for t in tasks
    ]


@router.get("/tasks/{task_id}")
async def get_task(task_id: str):
    db = _get_db()
    task = await db.get_task(task_id)
    if not task:
        raise HTTPException(404, "Задача не найдена")
    return {
        "task_id": task["id"],
        "mode": task["mode"],
        "status": task["status"],
        "total_urls": task["total_urls"],
        "processed_urls": task["processed_urls"] or 0,
        "found_contacts": task["found_contacts"] or 0,
        "errors_count": task["errors_count"] or 0,
        "created_at": task["created_at"],
        "output_file": task.get("output_file"),
    }


@router.get("/tasks/{task_id}/result")
async def download_result(task_id: str):
    db = _get_db()
    task = await db.get_task(task_id)
    if not task:
        raise HTTPException(404, "Задача не найдена")
    if task["status"] != "completed":
        raise HTTPException(400, "Задача ещё не завершена")
    output_file = task.get("output_file")
    if not output_file or not Path(output_file).exists():
        raise HTTPException(404, "Файл результата не найден")
    return FileResponse(
        path=output_file,
        filename=f"contacts_{task_id}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/tasks/{task_id}/pause")
async def pause_task(task_id: str):
    tm = _get_tm()
    await tm.pause_task(task_id)
    return {"status": "paused"}


@router.post("/tasks/{task_id}/resume")
async def resume_task(task_id: str):
    tm = _get_tm()
    await tm.resume_task(task_id)
    return {"status": "running"}


@router.post("/tasks/{task_id}/cancel")
async def cancel_task(task_id: str):
    tm = _get_tm()
    await tm.cancel_task(task_id)
    return {"status": "cancelled"}


# ═══ Blacklist ═══

@router.post("/blacklist/upload")
async def upload_blacklist(file: UploadFile = File(...)):
    db = _get_db()
    content = await file.read()
    added = 0
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if cell:
                    val = str(cell).strip().lower()
                    if "@" in val:
                        ok = await db.add_blacklist_entry("email", val)
                    elif "." in val:
                        ok = await db.add_blacklist_entry("domain", val)
                    else:
                        continue
                    if ok:
                        added += 1
        wb.close()
    except Exception:
        text = content.decode("utf-8", errors="ignore")
        for line in text.splitlines():
            val = line.strip().lower()
            if not val:
                continue
            if "@" in val:
                ok = await db.add_blacklist_entry("email", val)
            elif "." in val:
                ok = await db.add_blacklist_entry("domain", val)
            else:
                continue
            if ok:
                added += 1

    total = await db.count_blacklist()
    return {"entries_added": added, "total_entries": total}


@router.get("/blacklist")
async def get_blacklist():
    db = _get_db()
    return await db.get_blacklist()


@router.delete("/blacklist/{entry_id}")
async def delete_blacklist_entry(entry_id: int):
    db = _get_db()
    await db.delete_blacklist_entry(entry_id)
    return {"status": "deleted"}
