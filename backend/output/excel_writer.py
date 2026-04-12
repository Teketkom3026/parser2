"""Экспорт результатов в Excel."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.config import settings
from backend.utils.logger import get_logger

logger = get_logger("excel_writer")

# Стили
_HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_CELL_FONT = Font(name="Arial", size=10)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=True)

_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

_ALT_FILL = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

# Столбцы
COLUMNS = [
    ("site_url", "Сайт", 30),
    ("company_name", "Компания", 25),
    ("person_name", "ФИО", 25),
    ("position_raw", "Должность (исходная)", 25),
    ("position_norm", "Должность (норм.)", 25),
    ("role_category", "Категория должности", 18),
    ("person_email", "Email персоны", 28),
    ("email_type", "Тип email", 15),
    ("company_email", "Email компании", 28),
    ("person_phone", "Телефон персоны", 20),
    ("company_phone", "Телефон компании", 20),
    ("inn", "ИНН", 15),
    ("kpp", "КПП", 12),
    ("social_links", "Соцсети", 35),
    ("source_url", "Страница-источник", 35),
]


async def write_excel(contacts: list[dict], task_id: str) -> str:
    """Генерирует Excel-файл и возвращает путь."""
    results_dir = Path(settings.RESULTS_DIR)
    results_dir.mkdir(parents=True, exist_ok=True)

    output_path = results_dir / f"contacts_{task_id[:8]}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Контакты"

    # ═══════════ ЗАГОЛОВКИ ═══════════
    for col_idx, (key, label, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Закрепить заголовок
    ws.freeze_panes = "A2"

    # Автофильтр
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ═══════════ ДАННЫЕ ═══════════
    for row_idx, contact in enumerate(contacts, 2):
        is_alt = (row_idx % 2) == 0

        for col_idx, (key, label, width) in enumerate(COLUMNS, 1):
            value = contact.get(key, "")

            # Соцсети — объединяем список
            if key == "social_links" and isinstance(value, (list, tuple)):
                value = "\n".join(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.font = _CELL_FONT
            cell.alignment = _CELL_ALIGN
            cell.border = _BORDER

            if is_alt:
                cell.fill = _ALT_FILL

    # ═══════════ ЛИСТ СВОДКИ ═══════════
    ws_summary = wb.create_sheet("Сводка")
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 15

    summary_data = [
        ("Показатель", "Значение"),
        ("Всего контактов", len(contacts)),
        ("Уникальных сайтов", len(set(c.get("site_url", "") for c in contacts if c.get("site_url")))),
        ("С ФИО", sum(1 for c in contacts if c.get("person_name"))),
        ("С email персоны", sum(1 for c in contacts if c.get("person_email"))),
        ("С email компании", sum(1 for c in contacts if c.get("company_email"))),
        ("С телефоном", sum(1 for c in contacts if c.get("company_phone") or c.get("person_phone"))),
        ("С ИНН", sum(1 for c in contacts if c.get("inn"))),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws_summary.cell(row=row_idx, column=1, value=label)
        cell_b = ws_summary.cell(row=row_idx, column=2, value=value)

        if row_idx == 1:
            cell_a.font = _HEADER_FONT
            cell_a.fill = _HEADER_FILL
            cell_b.font = _HEADER_FONT
            cell_b.fill = _HEADER_FILL
        else:
            cell_a.font = _CELL_FONT
            cell_b.font = Font(name="Arial", size=10, bold=True)

        cell_a.border = _BORDER
        cell_b.border = _BORDER

    # Сохранение
    wb.save(str(output_path))
    logger.info("excel_written", path=str(output_path), rows=len(contacts))

    return str(output_path)
