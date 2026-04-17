"""Генерация итогового Excel-файла с листами по ролям."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from backend.utils.logger import get_logger

logger = get_logger("excel_generator")

# ── Заголовки колонок ──────────────────────────────────────────────────────────
HEADERS = [
    "Название компании",            # A
    "Сайт (домен)",                 # B
    "Общий email компании",         # C
    "Общий телефон компании",       # D
    "ФИО (полностью)",              # E
    "Фамилия",                      # F
    "Имя",                          # G
    "Отчество",                     # H
    "Должность (как на сайте)",     # I
    "Должность (нормализованная)",  # J
    "Категория (отрасль)",          # K
    "Личный email",                 # L
    "Личный телефон",               # M
    "ИНН",                          # N
    "КПП",                          # O
    "Соцсети",                      # P
    "URL страницы-источника",       # Q
    "Язык страницы",                # R
    "Дата сканирования",            # S
    "Статус",                       # T
    "Комментарий",                  # U
]

FIELD_MAP = [
    "company_name",
    "site_url",
    "company_email",
    "company_phone",
    "person_name",
    "last_name",
    "first_name",
    "patronymic",
    "position_raw",
    "position_norm",
    "role_category",
    "person_email",
    "person_phone",
    "inn",
    "kpp",
    "social_links",
    "page_url",
    "page_language",
    "scan_date",
    "status",
    "comment",
]

# ── Маппинг роль → лист ──────────────────────────────────────────────────────

# Ключевые слова в нормализованной должности для определения листа
SHEET_RULES = [
    ("Генеральные директора", [
        "генеральный директор", "гендиректор", "ceo", "президент компании",
        "управляющий директор", "исполнительный директор",
    ]),
    ("Финансовые директора", [
        "финансовый директор", "финдиректор", "cfo",
        "директор по финансам", "директор по экономике",
    ]),
    ("Главные бухгалтеры", [
        "главный бухгалтер", "главбух",
    ]),
    ("Главные инженеры", [
        "главный инженер", "технический директор", "cto",
        "директор по технологиям", "директор производства",
    ]),
]

_FALLBACK_SHEET = "Остальные"


def _get_sheet_for_contact(contact: dict) -> str:
    """Определить лист Excel для контакта по нормализованной должности."""
    pos_norm = (contact.get("position_norm") or "").lower()
    pos_raw = (contact.get("position_raw") or "").lower()
    combined = pos_norm + " " + pos_raw

    for sheet_name, keywords in SHEET_RULES:
        for kw in keywords:
            if kw in combined:
                return sheet_name
    return _FALLBACK_SHEET


# ── Мусорная фильтрация ────────────────────────────────────────────────────────

_JUNK_NAMES = {
    "Page Down", "Page Up", "Map Data", "Keyboard shortcuts",
    "Terms of Use", "Report a map error", "Scroll to zoom",
    "ФИО не найдено", "Satellite",
}


def _is_valid_contact(contact: dict) -> bool:
    name = (contact.get("person_name") or "").strip()
    pos = (contact.get("position_raw") or "").strip()
    email = (contact.get("person_email") or contact.get("company_email") or "").strip()
    phone = (contact.get("company_phone") or contact.get("person_phone") or "").strip()
    if name in _JUNK_NAMES:
        return False
    return bool(name) or bool(pos and (email or phone))


# ── Стили ─────────────────────────────────────────────────────────────────────

def _header_style(cell, bg_color: str = "3F51B5"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


_EVEN_FILL = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")

COL_WIDTHS = [
    24, 22, 26, 24,  # A-D
    30,              # E ФИО полностью
    18, 14, 18,      # F-H Фамилия Имя Отчество
    36, 30, 24,      # I-K Должность raw / norm / категория
    26, 22,          # L-M email / phone личные
    14, 12,          # N-O ИНН КПП
    30,              # P соцсети
    40,              # Q URL
    8, 14, 10, 24,   # R-U язык / дата / статус / комментарий
]

SHEET_COLORS = {
    "Генеральные директора":  "1A237E",
    "Финансовые директора":   "1B5E20",
    "Главные бухгалтеры":     "4A148C",
    "Главные инженеры":       "BF360C",
    _FALLBACK_SHEET:          "37474F",
}


def _write_sheet(ws, contacts: list[dict], header_color: str):
    """Записать данные на лист."""
    border = _thin_border()

    # Заголовки
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _header_style(cell, header_color)

    # Данные
    for row_idx, contact in enumerate(contacts, 2):
        for col_idx, field_name in enumerate(FIELD_MAP, 1):
            value = contact.get(field_name, "") or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = _EVEN_FILL

    # Ширины колонок
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    if len(contacts) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(contacts) + 1}"


def _dedup_contacts(contacts: list[dict]) -> list[dict]:
    """Дедупликация: убираем одинаковые (site_url + person_name + position_raw)."""
    seen: set = set()
    result = []
    for c in contacts:
        key = (
            (c.get("site_url") or "").strip().lower(),
            (c.get("person_name") or "").strip().lower(),
            (c.get("position_raw") or "").strip().lower(),
        )
        if key not in seen:
            seen.add(key)
            result.append(c)
    return result


def generate_excel(contacts: list[dict], output_path: str) -> str:
    # Фильтрация мусора
    contacts = [c for c in contacts if _is_valid_contact(c)]
    # Дедупликация
    contacts = _dedup_contacts(contacts)

    # Группировка по листам
    sheets: dict[str, list[dict]] = {
        "Генеральные директора":  [],
        "Финансовые директора":   [],
        "Главные бухгалтеры":     [],
        "Главные инженеры":       [],
        _FALLBACK_SHEET:          [],
    }
    for c in contacts:
        sheet_name = _get_sheet_for_contact(c)
        sheets[sheet_name].append(c)

    wb = Workbook()
    # Удаляем дефолтный лист
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Создаём листы в нужном порядке
    sheet_order = list(sheets.keys())
    for sheet_name in sheet_order:
        sheet_contacts = sheets[sheet_name]
        color = SHEET_COLORS.get(sheet_name, "37474F")
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, sheet_contacts, color)

    # ── Лист Все контакты ──────────────────────────────────────────────────
    ws_all = wb.create_sheet(title="Все контакты")
    _write_sheet(ws_all, contacts, "455A64")

    # ── Лист Сводка ───────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="Сводка")
    ws_sum.cell(1, 1, "Метрика").font = Font(bold=True)
    ws_sum.cell(1, 2, "Значение").font = Font(bold=True)

    summary_rows = [
        ("Всего записей",           len(contacts)),
        ("С ФИО",                   sum(1 for c in contacts if c.get("person_name"))),
        ("С личным email",          sum(1 for c in contacts if c.get("person_email"))),
        ("С должностью",            sum(1 for c in contacts if c.get("position_raw"))),
        ("С нормализованной должн.", sum(1 for c in contacts if c.get("position_norm"))),
        ("С ИНН",                   sum(1 for c in contacts if c.get("inn"))),
        ("Уникальных компаний",      len(set(c.get("site_url", "") for c in contacts))),
        ("", ""),
    ]
    # Разбивка по листам
    for sheet_name in sheet_order:
        count = len(sheets[sheet_name])
        summary_rows.append((f"Лист: {sheet_name}", count))

    for i, (label, val) in enumerate(summary_rows, 2):
        ws_sum.cell(i, 1, label)
        ws_sum.cell(i, 2, val)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15

    wb.save(output_path)
    logger.info("excel_generated", path=output_path, rows=len(contacts))
    return output_path
